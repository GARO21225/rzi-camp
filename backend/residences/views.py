
from rest_framework import viewsets, filters, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from accounts.permissions import TokenInQueryOrHeader
from .models import (InductionRecord, Batiment, Personnel, OccupationHistory, Demande,
    InductionCampConfig, InductionInfra, InductionRegle, InductionQuizQuestion)
from .serializers import (BatimentSerializer, PersonnelSerializer, OccupationHistorySerializer,
    DemandeSerializer, InductionRecordSerializer, InductionCampConfigSerializer,
    InductionInfraSerializer, InductionRegleSerializer, InductionQuizQuestionSerializer,
    InductionQuizQuestionPublicSerializer)
import csv, datetime
from django.http import HttpResponse
from natsort import natsorted

class PersonnelViewSet(viewsets.ModelViewSet):
    queryset = Personnel.objects.all()
    serializer_class = PersonnelSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ["nom","prenom","societe","numero"]

    def get_queryset(self):
        qs = Personnel.objects.all()
        t = self.request.query_params.get("type_personnel")
        if t: qs = qs.filter(type_personnel=t)
        return qs

    @action(detail=False, methods=['post'])
    def import_csv_data(self, request):
        """Import masse de personnel depuis données JSON"""
        import traceback
        from rest_framework.response import Response
        from django.db import connection
        rows = request.data.get('rows', [])
        ok = 0
        errors = []
        for i, row in enumerate(rows):
            nom = (row.get('nom') or '').strip().upper()
            prenom = (row.get('prenom') or '').strip().upper()
            societe = (row.get('societe') or 'N/A').strip().upper()
            email = (row.get('email') or '').strip()
            numero = (row.get('numero') or '').strip()
            type_p = (row.get('type_personnel') or 'roxgold').strip()
            if not nom or not prenom:
                errors.append(f"Ligne {i+2}: nom et prénom requis")
                continue
            try:
                # INSERT SQL direct pour éviter les triggers complexes
                with connection.cursor() as c:
                    c.execute(
                        """INSERT INTO residences_personnel
                        (nom, prenom, societe, email, numero, type_personnel,
                         actif, profil, qr_code_data, qr_code_string,
                         login_genere, password_genere, date_creation)
                        VALUES (%s,%s,%s,%s,%s,%s,TRUE,'agent','','',
                                '',''::text, NOW())
                        RETURNING id""",
                        [nom, prenom, societe, email, numero, type_p]
                    )
                    pid = c.fetchone()[0]
                # Générer QR et user via le modèle
                p = Personnel.objects.get(pk=pid)
                try:
                    p.generer_qr()
                    Personnel.objects.filter(pk=pid).update(
                        qr_code_data=p.qr_code_data,
                        qr_code_string=p.qr_code_string
                    )
                except Exception:
                    pass
                try:
                    p.creer_utilisateur()
                except Exception as e:
                    pass  # Personnel créé même si user échoue
                ok += 1
            except Exception as e:
                errors.append(f"Ligne {i+2}: {str(e)[:100]}")
        return Response({'imported': ok, 'errors': errors})

    def create(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser or (hasattr(request.user,"profile") and request.user.profile.role=="admin")):
            return Response({"error":"Seul l'admin peut créer du personnel."}, status=403)
        response = super().create(request, *args, **kwargs)
        p = Personnel.objects.get(id=response.data["id"])
        username, password = '', ''
        try:
            username, password = p.creer_utilisateur()
        except Exception as e:
            pass  # Ne pas bloquer si création user échoue
        data = dict(response.data)
        data["login_genere"] = username
        data["password_genere"] = password
        # Notifier uniquement sécurité, médical, agent d'accueil
        try:
            from evenements.models import SimpleNotification
            PROFILS_NOTIF = ['securite', 'medical', 'admin']
            destinataires = Personnel.objects.filter(
                profil__in=PROFILS_NOTIF, user__isnull=False
            ).select_related('user').exclude(user=request.user)[:20]
            for pers in destinataires:
                SimpleNotification.objects.create(
                    user=pers.user,
                    titre='👋 Nouveau personnel enregistré',
                    message=f"{p.prenom} {p.nom} ({p.societe or 'N/A'}) vient d'arriver. Préparez l'accueil et l'induction QHSE.",
                    type_notif='induction', lu=False
                )
        except Exception:
            pass
        return Response(data, status=201)


    @action(detail=False, methods=["post"])
    def regenerer_tous_qr(self, request):
        """Régénère les QR codes de tous les personnels (admin)"""
        user = request.user
        is_admin = user.is_staff or user.is_superuser or (hasattr(user,'profile') and user.profile.role=='admin')
        if not is_admin:
            return Response({"error":"Admin uniquement"}, status=403)
        count = 0
        for p in Personnel.objects.all():
            p.generer_qr()
            count += 1
        return Response({"ok":True,"regenerated":count,"message":f"{count} QR régénérés"})
    @action(detail=True, methods=["post"])
    def regenerer_qr(self, request, pk=None):
        p = self.get_object()
        p.qr_code_data = ""; p.qr_code_string = ""
        p.save(update_fields=["qr_code_data","qr_code_string"])
        p.generer_qr(); p.refresh_from_db()
        return Response(PersonnelSerializer(p).data)

    @action(detail=True, methods=["post"])
    def regenerer_compte(self, request, pk=None):
        if not request.user.is_staff:
            return Response({"error":"Admin uniquement"}, status=403)
        p = self.get_object()
        username, password = p.creer_utilisateur()
        return Response({"login":username,"password":password})

    @action(detail=True, methods=["patch"])
    def assigner_role(self, request, pk=None):
        user = request.user
        is_admin = user.is_staff or user.is_superuser or (hasattr(user,'profile') and user.profile.role=='admin')
        if not is_admin:
            return Response({"error":"Admin uniquement"}, status=403)
        p = self.get_object()
        role = request.data.get("role")
        valid_roles = ["admin","agent","restauration","technicien","menage"]
        if not role or role not in valid_roles:
            return Response({"error":f"Role invalide. Valeurs acceptées: {valid_roles}"}, status=400)
        from accounts.models import Profile
        if p.user:
            prof, _ = Profile.objects.get_or_create(user=p.user)
            prof.role = role
            prof.save(update_fields=["role"])
            if role == "admin":
                p.user.is_staff = True
                p.user.save(update_fields=["is_staff"])
        return Response({"ok":True,"role":role,"message":f"Role '{role}' attribué à {p.nom} {p.prenom}"})

    @action(detail=True, methods=["get"])
    def historique_voyages(self, request, pk=None):
        p = self.get_object()
        from voyages.models import Voyage
        voys = Voyage.objects.filter(personnel=p).order_by("-date_depart")
        return Response({
            "personnel": f"{p.nom} {p.prenom}",
            "societe": p.societe,
            "total_voyages": voys.count(),
            "en_voyage": voys.filter(statut="en_voyage").count(),
            "destinations_uniques": list(set(v.destination for v in voys if v.destination)),
            "voyages": [{
                "id":v.id, "destination":v.destination or "Non spécifiée",
                "motif":v.motif, "date_depart":str(v.date_depart),
                "date_retour_prevue":str(v.date_retour_prevue),
                "date_retour_effective":str(v.date_retour_effective) if v.date_retour_effective else None,
                "statut":v.get_statut_display() if hasattr(v,"get_statut_display") else v.statut,
                "chambre":v.batiment.residence if v.batiment else None,
            } for v in voys]
        })

    @action(detail=True, methods=["get"])
    def historique_chambres(self, request, pk=None):
        p = self.get_object()
        date_debut = request.query_params.get("date_debut")
        date_fin = request.query_params.get("date_fin")
        qs = OccupationHistory.objects.filter(personnel=p).select_related("batiment")
        if date_debut: qs = qs.filter(date_arrivee__gte=date_debut)
        if date_fin: qs = qs.filter(date_arrivee__lte=date_fin)
        return Response(OccupationHistorySerializer(qs, many=True).data)


    # ── Accès admin uniquement ──────────────────────────────────────────
    @staticmethod
    def _is_admin(user):
        """Vérifie si l'utilisateur est admin (is_staff OU profile.role='admin')"""
        if user.is_staff or user.is_superuser:
            return True
        try:
            return user.profile.role == 'admin'
        except Exception:
            return False

    def perform_create(self, serializer):
        demande = serializer.save(demandeur=self.request.user)
        # Email + SMS de bienvenue (asynchrone, ne bloque pas)
        try:
            from rzi_camp.notifications import envoyer_email_bienvenue, envoyer_sms_bienvenue
            import threading
            def send():
                envoyer_email_bienvenue(p)
                envoyer_sms_bienvenue(p)
            threading.Thread(target=send, daemon=True).start()
        except Exception:
            pass


    def destroy(self, request, *args, **kwargs):
        """Supprimer un personnel — admin uniquement (SQL direct pour éviter FK)"""
        if not self._is_admin(request.user):
            return Response({"error": "Admin requis"}, status=403)
        try:
            obj = self.get_object()
            personnel_info = f"{obj.nom} {obj.prenom}"
            pers_id = obj.id

            from django.db import connection
            with connection.cursor() as cursor:
                # 1. Tables liées voyages
                cursor.execute("DELETE FROM voyages_voyage WHERE personnel_id = %s", [pers_id])

                # 2. Tables liées restauration
                cursor.execute("DELETE FROM restauration_repaslog WHERE personnel_id = %s", [pers_id])

                # 3. Tables occupation history
                cursor.execute("DELETE FROM residences_occupationhistory WHERE personnel_id = %s", [pers_id])

                # 4. Batiments - SET NULL puis DELETE
                cursor.execute("UPDATE residences_batiment SET personnel_id = NULL WHERE personnel_id = %s", [pers_id])

                # 5. Demandes
                cursor.execute("SELECT user_id FROM residences_personnel WHERE id = %s", [pers_id])
                row = cursor.fetchone()
                if row and row[0]:
                    user_id = row[0]
                    cursor.execute("DELETE FROM residences_demande WHERE demandeur_id = %s", [user_id])

                # 6. Historical records
                try:
                    cursor.execute("DELETE FROM residences_historicalpersonnel WHERE personnel_ptr_id = %s", [pers_id])
                except: pass

                # 7. finally delete personnel
                cursor.execute("DELETE FROM residences_personnel WHERE id = %s", [pers_id])

            return Response({"ok": True, "message": f"Personnel supprimé: {personnel_info}"})
        except Exception as e:
            return Response({"error": f"Erreur: {str(e)}"}, status=400)

    def partial_update(self, request, *args, **kwargs):
        """Modifier un personnel — admin uniquement"""
        if not self._is_admin(request.user):
            return Response({"error": "Admin requis"}, status=403)
        return super().partial_update(request, *args, **kwargs)

    @action(detail=True, methods=["post"])
    def toggle_active(self, request, pk=None):
        """Activer/désactiver le compte User d'un Personnel"""
        if not self._is_admin(request.user):
            return Response({"error": "Admin requis"}, status=403)
        p = self.get_object()
        if not p.user:
            return Response({"error": "Ce personnel n'a pas de compte utilisateur"}, status=400)
        p.user.is_active = not p.user.is_active
        p.user.save(update_fields=["is_active"])
        return Response({
            "ok": True,
            "is_active": p.user.is_active,
            "message": f"Compte {'activé' if p.user.is_active else 'désactivé'}"
        })

    @action(detail=False, methods=["get"])
    def scans_historique(self, request):
        """Historique des scans de repas pour le restaurant"""
        from restauration.models import RepasLog
        from restauration.serializers import RepasLogSerializer
        type_repas = request.query_params.get("type_repas")
        date = request.query_params.get("date")
        qs = RepasLog.objects.select_related(
            "qr_token", "qr_token__personnel", "valide_par"
        ).all().order_by("-date_validation")
        if type_repas:
            qs = qs.filter(qr_token__type_repas=type_repas)
        if date:
            qs = qs.filter(date_validation__date=date)
        serializer = RepasLogSerializer(qs[:500], many=True)
        return Response({"count": len(serializer.data), "results": serializer.data})

    @action(detail=False, methods=["get"])
    def mon_profil(self, request):
        """Personnel lié à l'utilisateur connecté"""
        user = request.user
        p = None
        try:
            p = Personnel.objects.get(user=user)
        except Personnel.DoesNotExist:
            p = Personnel.objects.filter(login_genere=user.username).first()
        if not p and user.last_name:
            p = Personnel.objects.filter(nom__iexact=user.last_name.upper()).first()
        if not p:
            return Response({"detail": "Profil non trouvé"}, status=404)
        return Response(PersonnelSerializer(p).data)


    @action(detail=False, methods=["post"])
    def import_csv(self, request):
        """Importer une liste de personnel depuis CSV/Excel"""
        if not self._is_admin(request.user):
            return Response({"error": "Admin requis"}, status=403)

        file = request.FILES.get("file")
        if not file:
            return Response({"error": "Fichier requis (CSV ou Excel)"}, status=400)

        filename = file.name.lower()
        rows = []

        try:
            if filename.endswith(".csv"):
                import csv, codecs
                reader = csv.DictReader(codecs.iterdecode(file, "utf-8-sig"))
                rows = list(reader)
            elif filename.endswith((".xlsx", ".xls")):
                import openpyxl
                wb = openpyxl.load_workbook(file, read_only=True)
                ws = wb.active
                headers = [str(c.value or "").strip() for c in next(ws.iter_rows())]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
            else:
                return Response({"error": "Format non supporté. Utilisez .csv ou .xlsx"}, status=400)
        except Exception as e:
            return Response({"error": f"Erreur lecture fichier: {str(e)[:100]}"}, status=400)

        # Mapping colonnes flexibles
        COL_MAP = {
            "nom": ["nom","name","last_name","Nom","NOM"],
            "prenom": ["prenom","prénom","first_name","Prénom","PRENOM"],
            "societe": ["societe","société","company","Société","SOCIETE","entreprise"],
            "type_personnel": ["type","type_personnel","poste","Poste","TYPE"],
            "numero": ["numero","numéro","phone","tel","Téléphone","NUMERO"],
            "email": ["email","mail","Email","EMAIL"],
        }

        def get_val(row, keys):
            for k in keys:
                if k in row and row[k]:
                    return str(row[k]).strip()
            return ""

        created, skipped = 0, 0
        errors = []

        for i, row in enumerate(rows, 1):
            nom    = get_val(row, COL_MAP["nom"]).upper()
            prenom = get_val(row, COL_MAP["prenom"]).upper()
            if not nom or not prenom:
                skipped += 1
                errors.append(f"Ligne {i}: nom/prénom manquant")
                continue

            try:
                p, created_flag = Personnel.objects.get_or_create(
                    nom=nom, prenom=prenom,
                    defaults={
                        "societe":         get_val(row, COL_MAP["societe"]) or "ROXGOLD",
                        "type_personnel":  get_val(row, COL_MAP["type_personnel"]) or "Agent",
                        "numero":          get_val(row, COL_MAP["numero"]),
                        "email":           get_val(row, COL_MAP["email"]),
                    }
                )
                if created_flag:
                    # Générer QR
                    import qrcode, io, base64
                    qr_str = f"{p.pk:04d}"
                    p.qr_code_string = qr_str
                    qr_img = qrcode.make(qr_str, box_size=20, border=2)
                    buf = io.BytesIO()
                    qr_img.save(buf, format="PNG")
                    p.qr_code_data = base64.b64encode(buf.getvalue()).decode()
                    p.save(update_fields=["qr_code_string","qr_code_data"])
                    created += 1
                else:
                    skipped += 1
            except Exception as e:
                errors.append(f"Ligne {i}: {str(e)[:60]}")
                skipped += 1

        return Response({
            "created": created,
            "skipped": skipped,
            "total":   len(rows),
            "errors":  errors[:10],
            "message": f"{created} personnel(s) importé(s), {skipped} ignoré(s)"
        })

    @action(detail=True, methods=['patch'])
    def toggle_induction(self, request, pk=None):
        """Toggle induction_requise via SQL direct (champ ajouté via setup_db)"""
        from django.db import connection
        from rest_framework.response import Response
        valeur = request.data.get('induction_requise', True)
        try:
            with connection.cursor() as c:
                # Vérifier si colonne existe
                c.execute("SELECT EXISTS(SELECT FROM information_schema.columns WHERE table_name='residences_personnel' AND column_name='induction_requise')")
                col_exists = c.fetchone()[0]
                if not col_exists:
                    c.execute("ALTER TABLE residences_personnel ADD COLUMN induction_requise BOOLEAN NOT NULL DEFAULT TRUE")
                c.execute("UPDATE residences_personnel SET induction_requise=%s WHERE id=%s", [valeur, pk])
            return Response({'id': pk, 'induction_requise': valeur})
        except Exception as e:
            return Response({'detail': str(e)}, status=500)



class BatimentViewSet(viewsets.ModelViewSet):
    # IMPORTANT: keep queryset as QuerySet for get_object() to work
    queryset = Batiment.objects.select_related("personnel").all()
    serializer_class = BatimentSerializer
    filter_backends = []  # disable DRF filters, we do it manually

    def _build_qs(self, request=None):
        """Build filtered QuerySet — always returns a real QuerySet"""
        qs = Batiment.objects.select_related("personnel").all()
        params = (request or self.request).query_params
        statut = params.get("statut")
        bloc = params.get("bloc")
        residence = params.get("residence")
        futur_depart = params.get("futur_depart")
        search = params.get("search","")
        if statut: qs = qs.filter(statut=statut)
        if bloc: qs = qs.filter(bloc=bloc)
        if residence: qs = qs.filter(residence__icontains=residence)
        # Filter for own residence
        mon_residence = self.request.query_params.get("mon_residence")
        if mon_residence == "1":
            user = self.request.user
            if hasattr(user, "personnel"):
                qs = qs.filter(personnel=user.personnel)
        if futur_depart == "s1":
            today = datetime.date.today()
            qs = qs.filter(date_depart__gte=today, date_depart__lte=today+datetime.timedelta(days=7))
        if search:
            from django.db.models import Q
            qs = qs.filter(Q(residence__icontains=search)|Q(occupant__icontains=search)|Q(societe__icontains=search))
        return qs

    def get_queryset(self):
        # For single-object operations (retrieve, update, destroy), return full QuerySet
        return Batiment.objects.select_related("personnel").all()

    def list(self, request, *args, **kwargs):
        qs = self._build_qs(request)
        items = natsorted(list(qs), key=lambda x: x.residence)
        serializer = self.get_serializer(items, many=True)
        return Response({"count":len(items),"results":serializer.data})

    def partial_update(self, request, *args, **kwargs):
        # Use standard queryset for object lookup
        instance = Batiment.objects.get(pk=kwargs["pk"])
        old_personnel = instance.personnel
        data = request.data.copy()

        # Resolve personnel
        personnel_id = data.get("personnel")
        personnel_obj = None
        if personnel_id and str(personnel_id).strip() not in ("","null","None"):
            try:
                personnel_obj = Personnel.objects.get(pk=int(str(personnel_id)))
                if not data.get("occupant"):
                    data["occupant"] = f"{personnel_obj.nom} {personnel_obj.prenom}"
                if not data.get("societe"):
                    data["societe"] = personnel_obj.societe
            except (Personnel.DoesNotExist, ValueError, TypeError):
                data["personnel"] = None
        else:
            data["personnel"] = None

        serializer = self.get_serializer(instance, data=data, partial=True)
        serializer.is_valid(raise_exception=True)
        obj = serializer.save()

        today = datetime.date.today()
        if obj.statut == "Libre":
            Batiment.objects.filter(pk=obj.pk).update(
                personnel=None, occupant=None, societe=None,
                date_arrivee=None, date_depart=None
            )
            obj.refresh_from_db()
            if old_personnel:
                OccupationHistory.objects.filter(
                    batiment=obj, personnel=old_personnel, date_depart__isnull=True
                ).update(date_depart=today, motif_depart="Libération manuelle")

        elif obj.statut == "Occupé":
            p = obj.personnel or old_personnel
            # Only create history if explicitly confirmed (confirm=true param)
            confirm = request.data.get("confirm", "false")
            if p and obj.date_arrivee and str(confirm).lower() in ("true","1","yes"):
                OccupationHistory.objects.get_or_create(
                    batiment=obj, personnel=p, date_depart__isnull=True,
                    defaults={
                        "occupant_nom":obj.occupant or f"{p.nom} {p.prenom}",
                        "societe":obj.societe or p.societe,
                        "date_arrivee":obj.date_arrivee,
                        "enregistre_par":request.user,
                    }
                )

        return Response(self.get_serializer(obj).data)

    @action(detail=False, methods=["get"])
    def geojson(self, request):
        qs = self._build_qs(request)
        features = []
        for b in qs:
            # Utiliser geojson_geometry OU créer un Point depuis lat/lng
            geom = b.geojson_geometry
            if not geom and b.latitude and b.longitude:
                geom = {"type": "Point", "coordinates": [b.longitude, b.latitude]}
            if geom:
                p = b.personnel
                features.append({
                    "type":"Feature",
                    "properties":{
                        "id":b.id,"residence":b.residence,"bloc":b.bloc,"statut":b.statut,
                        "occupant":f"{p.nom} {p.prenom}" if p else b.occupant,
                        "societe":p.societe if p else b.societe,
                        "latitude":b.latitude,"longitude":b.longitude,
                    },
                    "geometry":b.geojson_geometry
                })
        return Response({"type":"FeatureCollection","features":features})

    @action(detail=False, methods=["get"])
    def stats(self, request):
        from django.db.models import Count
        qs = Batiment.objects.all()
        total = qs.count()
        par_statut = dict(qs.values_list("statut").annotate(n=Count("id")).values_list("statut","n"))
        par_bloc = list(qs.values("bloc").annotate(total=Count("id")).order_by("bloc"))
        today = datetime.date.today()
        departs_s1 = qs.filter(date_depart__gte=today, date_depart__lte=today+datetime.timedelta(days=7)).count()
        departs_s1_list = list(qs.filter(date_depart__gte=today, date_depart__lte=today+datetime.timedelta(days=7))
                               .select_related("personnel")
                               .values("residence","occupant","date_depart","personnel__nom","personnel__prenom"))
        return Response({
            "total":total,"par_statut":par_statut,"par_bloc":par_bloc,
            "taux_occupation":round(par_statut.get("Occupé",0)/total*100,1) if total else 0,
            "departs_s1":departs_s1,
            "departs_s1_list":departs_s1_list,
        })

    @action(detail=False, methods=["get"], permission_classes=[TokenInQueryOrHeader])
    def export_csv(self, request):
        qs = natsorted(list(Batiment.objects.select_related("personnel").all()), key=lambda x: x.residence)
        statut = request.query_params.get("statut")
        if statut: qs = [b for b in qs if b.statut==statut]
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = "attachment; filename=residences_rzi.csv"
        response.write("\ufeff")
        writer = csv.writer(response, delimiter=";")
        writer.writerow(["Residence","Bloc","Statut","Occupant","Societe","Type","Telephone","Date arrivee","Date depart"])
        for b in qs:
            p = b.personnel
            writer.writerow([b.residence,b.bloc,b.statut,
                f"{p.nom} {p.prenom}" if p else (b.occupant or ""),
                p.societe if p else (b.societe or ""),
                p.get_type_personnel_display() if p else "",
                p.numero if p else "",
                b.date_arrivee or "", b.date_depart or ""])
        return response

    @action(detail=False, methods=["get"], permission_classes=[TokenInQueryOrHeader])
    def export_par_bloc(self, request):
        from django.db.models import Count, Q
        qs = Batiment.objects.values("bloc").annotate(
            total=Count("id"), libres=Count("id",filter=Q(statut="Libre")),
            occupes=Count("id",filter=Q(statut="Occupé")),
            reserves=Count("id",filter=Q(statut="Réservé")),
            maintenance=Count("id",filter=Q(statut="Maintenance")),
        ).order_by("bloc")
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = "attachment; filename=rapport_blocs_rzi.csv"
        response.write("\ufeff")
        writer = csv.writer(response, delimiter=";")
        writer.writerow(["Bloc","Total","Libres","Occupes","Reserves","Maintenance","Taux %"])
        for r in qs:
            taux = round(r["occupes"]/r["total"]*100,1) if r["total"] else 0
            writer.writerow([r["bloc"],r["total"],r["libres"],r["occupes"],r["reserves"],r["maintenance"],taux])
        return response


    def perform_create(self, serializer):
        demande = serializer.save(demandeur=self.request.user)
        # Email + SMS de bienvenue (asynchrone, ne bloque pas)
        try:
            from rzi_camp.notifications import envoyer_email_bienvenue, envoyer_sms_bienvenue
            import threading
            def send():
                envoyer_email_bienvenue(p)
                envoyer_sms_bienvenue(p)
            threading.Thread(target=send, daemon=True).start()
        except Exception:
            pass


    def destroy(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser or (hasattr(request.user,"profile") and request.user.profile.role=="admin")):
            return Response({"error":"Suppression réservée à l'admin"}, status=403)
        return super().destroy(request, *args, **kwargs)
    @action(detail=False, methods=["get"])
    def mon_profil(self, request):
        """Retourne le Personnel lié à l'utilisateur connecté"""
        user = request.user
        # Chercher par user FK
        p = None
        try:
            p = Personnel.objects.get(user=user)
        except Personnel.DoesNotExist:
            pass
        
        # Fallback: chercher par login_genere
        if not p:
            p = Personnel.objects.filter(login_genere=user.username).first()
        
        # Fallback: chercher par nom/prenom
        if not p and user.last_name:
            p = Personnel.objects.filter(
                nom__iexact=user.last_name,
                prenom__iexact=user.first_name
            ).first()
        
        if not p:
            return Response({"detail": "Aucun profil personnel lié"}, status=404)
        
        return Response(PersonnelSerializer(p).data)


class OccupationHistoryViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = OccupationHistory.objects.select_related("batiment","personnel").all()
    serializer_class = OccupationHistorySerializer

    def get_queryset(self):
        qs = OccupationHistory.objects.select_related("batiment","personnel").all()
        batiment = self.request.query_params.get("batiment")
        personnel = self.request.query_params.get("personnel")
        date_debut = self.request.query_params.get("date_debut")
        date_fin = self.request.query_params.get("date_fin")
        if batiment: qs = qs.filter(batiment__residence__iexact=batiment)
        if personnel: qs = qs.filter(personnel_id=personnel)
        if date_debut: qs = qs.filter(date_arrivee__gte=date_debut)
        if date_fin: qs = qs.filter(date_arrivee__lte=date_fin)
        return qs


    @action(detail=False, methods=["get"], permission_classes=[TokenInQueryOrHeader])
    def export_csv(self, request):
        import csv
        from django.http import HttpResponse
        qs = OccupationHistory.objects.select_related("batiment","personnel").order_by("-date_arrivee")
        batiment = request.query_params.get("batiment")
        personnel = request.query_params.get("personnel")
        if batiment: qs = qs.filter(batiment__residence__icontains=batiment)
        if personnel: qs = qs.filter(personnel_id=personnel)
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = "attachment; filename=historique_occupation.csv"
        response.write("\ufeff")
        writer = csv.writer(response, delimiter=";")
        writer.writerow(["Residence","Bloc","Occupant","Societe","Arrivee","Depart","Duree jours","Motif depart"])
        import datetime
        today = datetime.date.today()
        for h in qs:
            d1 = h.date_arrivee
            d2 = h.date_depart or today
            days = (d2-d1).days
            writer.writerow([h.batiment.residence,h.batiment.bloc,h.occupant_nom,h.societe or "",
                str(h.date_arrivee),str(h.date_depart) if h.date_depart else "En cours",days,h.motif_depart or ""])
        return response

    @action(detail=False, methods=["get"])
    def recherche(self, request):
        """
        Recherche croisée historisation :
        - Toutes chambres si pas de filtre batiment
        - Filtre par résidence (dropdown ou saisie libre)
        - Filtre par nom occupant (saisie libre)
        - Filtre par personnel (dropdown)
        - Filtre par période
        """
        batiment_q = request.query_params.get("batiment","").strip()
        personnel_id = request.query_params.get("personnel","").strip()
        date_debut = request.query_params.get("date_debut","").strip()
        date_fin = request.query_params.get("date_fin","").strip()
        nom_search = request.query_params.get("nom","").strip()

        from django.db.models import Q

        # Build history queryset — no mandatory filter, show all if no criteria
        qs = OccupationHistory.objects.select_related("batiment","personnel").all()
        if batiment_q:
            qs = qs.filter(batiment__residence__icontains=batiment_q)
        if personnel_id:
            qs = qs.filter(personnel_id=personnel_id)
        if nom_search:
            qs = qs.filter(
                Q(occupant_nom__icontains=nom_search) |
                Q(personnel__nom__icontains=nom_search) |
                Q(personnel__prenom__icontains=nom_search) |
                Q(batiment__residence__icontains=nom_search)
            )
        if date_debut:
            qs = qs.filter(Q(date_depart__gte=date_debut)|Q(date_depart__isnull=True))
        if date_fin:
            qs = qs.filter(date_arrivee__lte=date_fin)

        today = datetime.date.today()
        results = []

        # Also include current batiment occupants not yet in history
        bats_qs = Batiment.objects.select_related("personnel").filter(statut="Occupé")
        if batiment_q:
            bats_qs = bats_qs.filter(residence__icontains=batiment_q)
        if personnel_id:
            bats_qs = bats_qs.filter(personnel_id=personnel_id)
        if nom_search:
            bats_qs = bats_qs.filter(
                Q(occupant__icontains=nom_search) |
                Q(personnel__nom__icontains=nom_search) |
                Q(personnel__prenom__icontains=nom_search) |
                Q(residence__icontains=nom_search)
            )

        existing_bat_ids = set(qs.values_list("batiment_id", flat=True))
        for b in bats_qs:
            if b.id not in existing_bat_ids and b.date_arrivee:
                d1 = b.date_arrivee
                days = (today - d1).days
                results.append({
                    "id": f"bat-{b.id}",
                    "residence": b.residence,
                    "bloc": b.bloc,
                    "occupant": b.occupant or (f"{b.personnel.nom} {b.personnel.prenom}" if b.personnel else "—"),
                    "societe": b.societe or (b.personnel.societe if b.personnel else ""),
                    "personnel_id": b.personnel_id,
                    "date_arrivee": str(b.date_arrivee),
                    "date_depart": str(b.date_depart) if b.date_depart else None,
                    "duree_jours": max(days, 1),
                    "en_cours": True,
                    "motif_depart": "",
                    "source": "current",
                })

        for h in qs.order_by("-date_arrivee")[:500]:
            d1 = h.date_arrivee
            d2 = h.date_depart or today
            days = (d2 - d1).days
            results.append({
                "id": h.id,
                "residence": h.batiment.residence,
                "bloc": h.batiment.bloc,
                "occupant": h.occupant_nom,
                "societe": h.societe,
                "personnel_id": h.personnel_id,
                "date_arrivee": str(h.date_arrivee),
                "date_depart": str(h.date_depart) if h.date_depart else None,
                "duree_jours": max(days, 1),
                "en_cours": h.date_depart is None,
                "motif_depart": h.motif_depart,
                "source": "history",
            })

        results.sort(key=lambda x: x["date_arrivee"], reverse=True)
        return Response({"count":len(results),"results":results})



class OccupationHistoryAdminViewSet(viewsets.ModelViewSet):
    """Admin-only: correct or delete wrong history entries"""
    queryset = OccupationHistory.objects.select_related("batiment","personnel").all()
    serializer_class = OccupationHistorySerializer

    def perform_create(self, serializer):
        demande = serializer.save(demandeur=self.request.user)
        # Email + SMS de bienvenue (asynchrone, ne bloque pas)
        try:
            from rzi_camp.notifications import envoyer_email_bienvenue, envoyer_sms_bienvenue
            import threading
            def send():
                envoyer_email_bienvenue(p)
                envoyer_sms_bienvenue(p)
            threading.Thread(target=send, daemon=True).start()
        except Exception:
            pass


    def destroy(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser or (hasattr(request.user,"profile") and request.user.profile.role=="admin")):
            return Response({"error":"Admin uniquement"}, status=403)
        instance = self.get_object()
        batiment = instance.batiment
        # If this was the current occupant, do NOT free the room — just remove the wrong history
        instance.delete()
        return Response({"ok":True,"message":"Entrée historique supprimée (chambre non modifiée)"})

    def update(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser or (hasattr(request.user,"profile") and request.user.profile.role=="admin")):
            return Response({"error":"Admin uniquement"}, status=403)
        return super().partial_update(request, *args, **kwargs)



from rest_framework import viewsets, filters
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.utils import timezone
from django.contrib.auth.models import User
from .models import InductionRecord, Demande, Batiment, Personnel
from .serializers import DemandeSerializer

class DemandeViewSet(viewsets.ModelViewSet):
    serializer_class = DemandeSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ["demandeur__first_name","demandeur__last_name","residence_souhaitee"]

    def get_queryset(self):
        user = self.request.user
        is_admin = user.is_staff or user.is_superuser
        if is_admin:
            qs = Demande.objects.select_related("demandeur","traite_par").all()
        else:
            qs = Demande.objects.filter(demandeur=user)
        
        statut = self.request.query_params.get("statut")
        type_d = self.request.query_params.get("type_demande")
        if statut: qs = qs.filter(statut=statut)
        if type_d: qs = qs.filter(type_demande=type_d)
        return qs

    def perform_create(self, serializer):
        demande = serializer.save(demandeur=self.request.user)
        # Notify admins
        try:
            demande.notifier_admin()
        except Exception:
            pass

    def perform_create(self, serializer):
        demande = serializer.save(demandeur=self.request.user)
        # Email + SMS de bienvenue (asynchrone, ne bloque pas)
        try:
            from rzi_camp.notifications import envoyer_email_bienvenue, envoyer_sms_bienvenue
            import threading
            def send():
                envoyer_email_bienvenue(p)
                envoyer_sms_bienvenue(p)
            threading.Thread(target=send, daemon=True).start()
        except Exception:
            pass


    def destroy(self, request, *args, **kwargs):
        """Only admin can hard-delete; agents can only cancel"""
        instance = self.get_object()
        if request.user.is_staff or request.user.is_superuser:
            return super().destroy(request, *args, **kwargs)
        return Response({"error":"Admin uniquement pour suppression"}, status=403)

    # ── Admin actions ──
    @action(detail=True, methods=["post"])
    def valider(self, request, pk=None):
        """Admin valide la demande"""
        if not (request.user.is_staff or request.user.is_superuser or (hasattr(request.user,"profile") and request.user.profile.role=="admin")):
            return Response({"error":"Admin uniquement"}, status=403)
        demande = self.get_object()
        if demande.statut not in ("en_attente", "proposition"):
            return Response({"error":"Demande non en attente"}, status=400)
        
        demande.statut = "validee"
        demande.traite_par = request.user
        demande.commentaire_admin = request.data.get("commentaire","Demande validée")
        demande.date_traitement = timezone.now()
        
        # Execute the action
        if demande.type_demande == "reservation_residence":
            residence = request.data.get("residence_attribuee") or demande.residence_souhaitee or demande.proposition_admin.get("residence")
            if residence:
                demande.residence_attribuee = residence
                try:
                    bat = Batiment.objects.get(residence=residence)
                    # Assign if libre
                    if bat.statut == "Libre":
                        bat.statut = "Réservé"
                        bat.occupant = demande.demandeur.get_full_name()
                        bat.date_arrivee = demande.date_debut_souhaitee
                        bat.date_depart = demande.date_fin_souhaitee
                        bat.save()
                except Batiment.DoesNotExist:
                    pass
        
        elif demande.type_demande == "voyage":
            from voyages.models import Voyage
            data = demande.donnees
            if demande.demandeur:
                p = getattr(demande.demandeur,"personnel",None)
                if p:
                    import datetime
                    try:
                        Voyage.objects.create(
                            personnel=p,
                            destination=data.get("destination",""),
                            motif=data.get("motif",""),
                            date_depart=demande.date_debut_souhaitee or datetime.date.today(),
                            date_retour_prevue=demande.date_fin_souhaitee or (demande.date_debut_souhaitee or datetime.date.today()) + datetime.timedelta(days=7),
                            enregistre_par=request.user,
                        )
                    except Exception as ve:
                        pass  # Continue even if voyage creation fails
        
        demande.save()
        try:
            demande.notifier_demandeur(f"✅ Votre demande a été validée.\n{demande.commentaire_admin}")
        except Exception:
            pass
        return Response(DemandeSerializer(demande).data)

    @action(detail=True, methods=["post"])
    def rejeter(self, request, pk=None):
        """Admin rejette la demande"""
        if not (request.user.is_staff or request.user.is_superuser or (hasattr(request.user,"profile") and request.user.profile.role=="admin")):
            return Response({"error":"Admin uniquement"}, status=403)
        demande = self.get_object()
        demande.statut = "rejetee"
        demande.traite_par = request.user
        demande.commentaire_admin = request.data.get("commentaire","Demande rejetée")
        demande.date_traitement = timezone.now()
        demande.save()
        try:
            demande.notifier_demandeur(f"❌ Votre demande a été rejetée.\nMotif: {demande.commentaire_admin}")
        except Exception:
            pass
        return Response(DemandeSerializer(demande).data)

    @action(detail=True, methods=["post"])
    def proposer(self, request, pk=None):
        """Admin fait une contre-proposition"""
        if not (request.user.is_staff or request.user.is_superuser or (hasattr(request.user,"profile") and request.user.profile.role=="admin")):
            return Response({"error":"Admin uniquement"}, status=403)
        demande = self.get_object()
        demande.statut = "proposition"
        demande.traite_par = request.user
        demande.commentaire_admin = request.data.get("commentaire","Voir la proposition ci-dessous")
        demande.proposition_admin = request.data.get("proposition",{})
        demande.date_traitement = timezone.now()
        demande.save()
        
        prop_str = "\n".join([f"• {k}: {v}" for k,v in demande.proposition_admin.items()])
        try:
            demande.notifier_demandeur(f"💬 Proposition de l'admin:\n{prop_str}\n\n{demande.commentaire_admin}")
        except Exception:
            pass
        return Response(DemandeSerializer(demande).data)

    # ── Demandeur actions ──
    @action(detail=True, methods=["post"])
    def accepter_proposition(self, request, pk=None):
        """Demandeur accepte la proposition de l'admin"""
        demande = self.get_object()
        if demande.demandeur != request.user:
            return Response({"error":"Non autorisé"}, status=403)
        if demande.statut != "proposition":
            return Response({"error":"Pas de proposition en attente"}, status=400)
        # Validate with proposition
        demande.statut = "acceptee"
        demande.date_reponse = timezone.now()
        demande.save()
        return Response(DemandeSerializer(demande).data)

    @action(detail=True, methods=["post"])
    def refuser_proposition(self, request, pk=None):
        """Demandeur refuse la proposition"""
        demande = self.get_object()
        if demande.demandeur != request.user:
            return Response({"error":"Non autorisé"}, status=403)
        if demande.statut != "proposition":
            return Response({"error":"Pas de proposition"}, status=400)
        demande.statut = "rejetee"
        demande.date_reponse = timezone.now()
        demande.save()
        return Response(DemandeSerializer(demande).data)

    @action(detail=True, methods=["post"])
    def annuler(self, request, pk=None):
        """Demandeur annule sa propre demande"""
        demande = self.get_object()
        if demande.demandeur != request.user and not request.user.is_staff:
            return Response({"error":"Non autorisé"}, status=403)
        if demande.statut in ("validee","rejetee"):
            return Response({"error":"Demande déjà traitée"}, status=400)
        demande.statut = "annulee"
        demande.date_reponse = timezone.now()
        demande.save()
        return Response(DemandeSerializer(demande).data)

    @action(detail=False, methods=["get"])
    def stats(self, request):
        """Stats pour l'admin"""
        if not (request.user.is_staff or request.user.is_superuser or (hasattr(request.user,"profile") and request.user.profile.role=="admin")):
            return Response({"error":"Admin uniquement"}, status=403)
        from django.db.models import Count
        qs = Demande.objects.all()
        return Response({
            "total": qs.count(),
            "en_attente": qs.filter(statut="en_attente").count(),
            "validees": qs.filter(statut="validee").count(),
            "rejetees": qs.filter(statut="rejetee").count(),
            "propositions": qs.filter(statut="proposition").count(),
            "par_type": dict(qs.values_list("type_demande").annotate(n=Count("id")).values_list("type_demande","n")),
        })
    @action(detail=False, methods=['post'])
    def desactiver_expires(self, request):
        """Désactiver automatiquement le personnel temporaire expiré"""
        from django.utils import timezone
        expirés = Personnel.objects.filter(
            est_temporaire=True,
            date_expiration__lt=timezone.now(),
            actif=True
        )
        count = expirés.count()
        expirés.update(actif=False)
        return Response({'desactives': count, 'message': f'{count} compte(s) temporaire(s) désactivé(s)'})



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def declarer_soustraitants_masse(request):
    """Créer N sous-traitants temporaires pour une société"""
    societe = request.data.get('societe','').strip()
    nombre  = int(request.data.get('nombre', 0))
    duree_h = int(request.data.get('duree_h', 72))
    if not societe: return Response({'error':'Société requise'}, status=400)
    if not (1 <= nombre <= 100): return Response({'error':'Nombre invalide (1-100)'}, status=400)

    import random, string, datetime
    from django.utils import timezone
    from django.contrib.auth.models import User

    prefix  = ''.join(c for c in societe.upper() if c.isalpha())[:3] or 'ST'
    expire  = timezone.now() + datetime.timedelta(hours=duree_h)
    created = []

    for i in range(1, nombre+1):
        login = f"{prefix}{i:02d}"
        cnt   = 1
        while User.objects.filter(username=login).exists():
            login = f"{prefix}{i:02d}_{cnt}"; cnt += 1
        pwd  = ''.join(random.choices(string.ascii_letters+string.digits, k=8))
        user = User.objects.create_user(username=login, password=pwd,
            first_name=societe, last_name=f'ST-{i:02d}', is_active=True)
        try:
            from accounts.models import Profile
            Profile.objects.get_or_create(user=user, defaults={'role':'agent'})
        except: pass
        p = Personnel.objects.create(
            nom=societe, prenom=f'ST-{i:02d}', societe=societe,
            numero='', type_personnel='sous_traitant', user=user, actif=True)
        try: p.generer_qr()
        except: pass
        created.append({'id':p.id,'login':login,'pwd':pwd,
            'nom':f"{societe} ST-{i:02d}",'expire':expire.strftime('%d/%m/%Y %H:%M')})

    return Response({'ok':True,'created':len(created),'societe':societe,
        'expire':expire.strftime('%d/%m/%Y %H:%M'),'duree_h':duree_h,
        'agents':created,'message':f"{len(created)} sous-traitants {societe} créés — accès {duree_h}h"})


# ── Induction QHSE ───────────────────────────────────────
from rest_framework import serializers as drf_ser
class InductionRecordSerializer(drf_ser.ModelSerializer):
    personnel_nom = drf_ser.SerializerMethodField()
    progression   = drf_ser.SerializerMethodField()

    def get_personnel_nom(self, obj):
        return f"{obj.personnel.nom} {obj.personnel.prenom}"

    def get_progression(self, obj):
        return obj.progression_pct()

    class Meta:
        model  = InductionRecord
        fields = '__all__'



    @action(detail=False, methods=['get'])
    def expirant_bientot(self, request):
        """Personnel dont l'induction expire dans les 30 jours (valide depuis >11 mois)."""
        from django.db import connection
        from rest_framework.response import Response
        from datetime import datetime, timedelta
        seuil = datetime.now() - timedelta(days=335)  # 11 mois = 335 jours
        try:
            with connection.cursor() as c:
                c.execute(
                    "SELECT ir.id, ir.personnel_id, p.nom, p.prenom, p.societe, ir.mis_a_jour "
                    "FROM residences_inductionrecord ir "
                    "JOIN residences_personnel p ON p.id=ir.personnel_id "
                    "WHERE ir.statut='valide' AND ir.mis_a_jour<%s "
                    "ORDER BY ir.mis_a_jour ASC LIMIT 50", [seuil])
                rows = c.fetchall()
                return Response([{
                    'id':r[0],'personnel_id':r[1],'nom':r[2],
                    'prenom':r[3],'societe':r[4],'depuis':str(r[5])
                } for r in rows])
        except Exception as e:
            return Response({'error': str(e)}, status=200)

class InductionRecordViewSet(viewsets.ModelViewSet):
    queryset           = InductionRecord.objects.select_related('personnel').all()
    serializer_class   = InductionRecordSerializer
    permission_classes = [IsAuthenticated]

    def destroy(self, request, *args, **kwargs):
        """Suppression via SQL direct pour éviter les erreurs de migration."""
        from django.db import connection
        from rest_framework.response import Response
        pk = kwargs.get('pk')
        try:
            with connection.cursor() as c:
                c.execute("SELECT EXISTS(SELECT FROM information_schema.tables WHERE table_name='residences_inductionrecord')")
                if c.fetchone()[0]:
                    c.execute("DELETE FROM residences_inductionrecord WHERE id=%s", [pk])
            return Response(status=204)
        except Exception as e:
            # Fallback: suppression ORM
            try:
                return super().destroy(request, *args, **kwargs)
            except Exception:
                return Response({'detail': str(e)}, status=200)  # 200 pour éviter crash frontend

    def get_permissions(self):
        """Toutes les actions nécessitent une authentification — un dossier d'induction
        contient des données personnelles (nom, société, scores) qui ne doivent pas être
        lisibles ni modifiables sans session valide."""
        return [IsAuthenticated()]

    def get_queryset(self):
        qs = super().get_queryset()
        pid = self.request.query_params.get('personnel')
        if pid: qs = qs.filter(personnel_id=pid)
        return qs

    @action(detail=False, methods=['post'])
    def update_etape(self, request):
        """Mettre a jour une etape pour un personnel."""
        personnel_id  = request.data.get('personnel_id')
        etape_key     = request.data.get('etape')
        done          = request.data.get('done', True)
        extra_data    = request.data.get('data', {})
        field_to_save = request.data.get('field')  # form_data|docs_data|medical_data|quiz_score

        if not personnel_id:
            return Response({'error': 'personnel_id requis'}, status=400)
        if not etape_key:
            return Response({'error': 'etape requise'}, status=400)

        try:
            from residences.models import Personnel, InductionRecord
            from django.utils import timezone
            import datetime
            
            personnel = Personnel.objects.get(id=personnel_id)
            record, created = InductionRecord.objects.get_or_create(personnel=personnel)
            
            # Mettre a jour les etapes (gerer None)
            etapes = dict(record.etapes_data or {})
            etapes[etape_key] = {
                'done': done, 
                'date': timezone.now().isoformat(),
                **(extra_data if isinstance(extra_data, dict) else {})
            }
            record.etapes_data = etapes
            
            # Sauvegarder donnees specifiques
            if field_to_save == 'form_data' and isinstance(extra_data, dict):
                record.form_data = {**(record.form_data or {}), **extra_data}
            elif field_to_save == 'docs_data' and isinstance(extra_data, dict):
                record.docs_data = {**(record.docs_data or {}), **extra_data}
            elif field_to_save == 'medical_data' and isinstance(extra_data, dict):
                record.medical_data = {**(record.medical_data or {}), **extra_data}
            elif field_to_save == 'quiz_score':
                score = extra_data.get('score') if isinstance(extra_data, dict) else None
                if score is not None:
                    record.quiz_score = score
                record.quiz_tentatives = (record.quiz_tentatives or 0) + 1
                
            # Verifier si induction complete
            ETAPES_REQUISES = ['accueil','documents','formation','quiz','medical','badge']
            if all(etapes.get(e, {}).get('done') for e in ETAPES_REQUISES):
                record.statut = 'valide'
                record.date_fin = timezone.now()
                record.badge_emis = True
                record.badge_date = timezone.now()
                record.badge_expire = (timezone.now() + datetime.timedelta(days=365)).date()
                
            record.save()
            
            return Response({
                'ok': True,
                'statut': record.statut,
                'progression': record.progression_pct(),
                'etapes': record.etapes_data,
                'created': created,
            })
        except Personnel.DoesNotExist:
            return Response({'error': f'Personnel {personnel_id} introuvable'}, status=404)
        except Exception as e:
            import traceback
            return Response({
                'error': str(e), 
                'traceback': traceback.format_exc()
            }, status=500)


# ═══════════════════════════════════════════════════════════════════
#  CONTENU ÉDITABLE — Induction Camp
#  CRUD complet réservé aux admins en écriture ; lecture ouverte à tout
#  utilisateur authentifié (le personnel doit pouvoir consulter le
#  contenu de son propre parcours d'induction).
# ═══════════════════════════════════════════════════════════════════

def _ensure_induction_tables():
    """Crée les 4 tables de contenu Induction Camp si elles n'existent pas encore.
    Auto-réparation sans nécessiter le secret SETUP_DB_SECRET — ces CREATE TABLE
    sont strictement additifs et sûrs à exécuter de façon répétée (IF NOT EXISTS).
    Évite que l'admin voie un onglet vide silencieusement si /api/setup-db/ n'a
    jamais été appelé après le déploiement de ces modèles."""
    from django.db import connection
    try:
        with connection.cursor() as c:
            c.execute("""CREATE TABLE IF NOT EXISTS residences_inductioncampconfig (
                id BIGSERIAL PRIMARY KEY,
                nom VARCHAR(200) NOT NULL DEFAULT 'Camp Résidentiel',
                site VARCHAR(200) NOT NULL DEFAULT '',
                capacite INTEGER NOT NULL DEFAULT 0,
                superficie VARCHAR(50) NOT NULL DEFAULT '',
                altitude VARCHAR(50) NOT NULL DEFAULT '',
                duree_parcours_min INTEGER NOT NULL DEFAULT 15,
                date_maj TIMESTAMP WITH TIME ZONE NOT NULL DEFAULT NOW()
            )""")
            c.execute("""CREATE TABLE IF NOT EXISTS residences_inductioninfra (
                id BIGSERIAL PRIMARY KEY,
                titre VARCHAR(100) NOT NULL DEFAULT '',
                emoji VARCHAR(10) NOT NULL DEFAULT '🏠',
                couleur VARCHAR(20) NOT NULL DEFAULT '#3b82f6',
                description TEXT NOT NULL DEFAULT '',
                details JSONB NOT NULL DEFAULT '[]',
                photo_base64 TEXT NOT NULL DEFAULT '',
                photo_mime VARCHAR(50) NOT NULL DEFAULT 'image/jpeg',
                ordre INTEGER NOT NULL DEFAULT 0,
                actif BOOLEAN NOT NULL DEFAULT TRUE,
                date_maj TIMESTAMP WITH TIME ZONE NOT NULL DEFAULT NOW()
            )""")
            c.execute("""CREATE TABLE IF NOT EXISTS residences_inductionregle (
                id BIGSERIAL PRIMARY KEY,
                titre VARCHAR(150) NOT NULL DEFAULT '',
                emoji VARCHAR(10) NOT NULL DEFAULT '📋',
                niveau VARCHAR(20) NOT NULL DEFAULT 'standard',
                texte TEXT NOT NULL DEFAULT '',
                ordre INTEGER NOT NULL DEFAULT 0,
                actif BOOLEAN NOT NULL DEFAULT TRUE,
                date_maj TIMESTAMP WITH TIME ZONE NOT NULL DEFAULT NOW()
            )""")
            c.execute("""CREATE TABLE IF NOT EXISTS residences_inductionquizquestion (
                id BIGSERIAL PRIMARY KEY,
                question TEXT NOT NULL DEFAULT '',
                options JSONB NOT NULL DEFAULT '[]',
                bonne_reponse SMALLINT NOT NULL DEFAULT 0,
                explication TEXT NOT NULL DEFAULT '',
                ordre INTEGER NOT NULL DEFAULT 0,
                actif BOOLEAN NOT NULL DEFAULT TRUE,
                date_maj TIMESTAMP WITH TIME ZONE NOT NULL DEFAULT NOW()
            )""")
        return True
    except Exception:
        return False


def _is_admin_user(user):
    if user.is_staff or user.is_superuser:
        return True
    try:
        return user.profile.role == 'admin'
    except Exception:
        return False


class InductionAdminWriteMixin:
    """Lecture pour tout utilisateur authentifié, écriture réservée aux admins."""
    def get_permissions(self):
        return [IsAuthenticated()]

    def list(self, request, *args, **kwargs):
        """Auto-répare la table si elle n'existe pas encore (cas d'un déploiement
        où /api/setup-db/ n'a pas été appelé manuellement) plutôt que de renvoyer
        une 500 qui laisse l'onglet admin vide sans explication.
        Important : ne catche QUE l'erreur 'table manquante' — toute autre erreur
        remonte normalement, pour ne jamais masquer un vrai bug derrière un faux
        'aucune donnée' silencieux."""
        try:
            return super().list(request, *args, **kwargs)
        except Exception as e:
            msg = str(e).lower()
            table_missing = 'does not exist' in msg or 'no such table' in msg
            if not table_missing:
                raise  # erreur réelle (pas liée à une table absente) : ne pas la masquer
            if _ensure_induction_tables():
                return super().list(request, *args, **kwargs)
            raise

    def _check_admin(self, request):
        if not _is_admin_user(request.user):
            return Response({"error": "Admin uniquement"}, status=403)
        return None

    def create(self, request, *args, **kwargs):
        err = self._check_admin(request)
        if err: return err
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        err = self._check_admin(request)
        if err: return err
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        err = self._check_admin(request)
        if err: return err
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        err = self._check_admin(request)
        if err: return err
        return super().destroy(request, *args, **kwargs)


class InductionCampConfigViewSet(InductionAdminWriteMixin, viewsets.ModelViewSet):
    queryset = InductionCampConfig.objects.all()
    serializer_class = InductionCampConfigSerializer

    @action(detail=False, methods=["post"])
    def importer_donnees_originales(self, request):
        """Importe une seule fois le contenu qui était codé en dur dans
        InductionCamp.jsx avant la refonte admin (8 infrastructures, 8 règles,
        5 questions de quiz, config du camp). Idempotent : si des infrastructures
        existent déjà, ne réimporte rien pour éviter les doublons — un admin qui
        a déjà commencé à personnaliser le contenu ne perd jamais son travail.
        Réservé aux admins (même garde que les autres écritures)."""
        if not _is_admin_user(request.user):
            return Response({"error": "Admin uniquement"}, status=403)

        _ensure_induction_tables()

        if InductionInfra.objects.exists() or InductionRegle.objects.exists() or InductionQuizQuestion.objects.exists():
            return Response({
                "status": "skipped",
                "message": "Du contenu existe déjà — import ignoré pour ne pas créer de doublons. "
                           "Supprimez tout le contenu existant si vous voulez réimporter depuis zéro.",
            })

        if not InductionCampConfig.objects.exists():
            InductionCampConfig.objects.create(
                nom="Camp Résidentiel Roxgold Sango",
                site="Mine d'Or de Sango · Côte d'Ivoire",
                capacite=247, superficie="12 ha", altitude="347m",
                duree_parcours_min=15,
            )

        infras_data = [
            ("Résidences", "🏠", "#3b82f6",
             "5 blocs résidentiels A–E + bloc VIP. Chambres individuelles climatisées, salle de bain privée, Wi-Fi haut débit.",
             ["Clim réglable 20–26°C", "Wi-Fi 50 Mbps", "Linge de lit fourni", "Ménage quotidien"]),
            ("Restauration", "🍽️", "#f59e0b",
             "Cafétéria principale ouverte 6h–21h. 3 repas par jour inclus. Régimes spéciaux disponibles.",
             ["Buffet petit-déjeuner 6h–9h", "Déjeuner 11h30–13h30", "Dîner 18h–21h", "Snacks disponibles 24h"]),
            ("Infirmerie", "🏥", "#ef4444",
             "Infirmière présente 24h/24. Médecin en visite 3×/semaine. Évacuation médicale disponible.",
             ["Urgences 24h/7j", "Médicaments de base fournis", "Évacuation hélico si nécessaire", "Formulaires médicaux en ligne"]),
            ("Sport & Loisirs", "⚽", "#10b981",
             "Terrain de foot, salle de muscu, basket, ping-pong. Horaires : 6h–8h et 17h–20h.",
             ["Terrain foot éclairé", "Salle musculation équipée", "Court basketball", "Salle TV & bibliothèque"]),
            ("Laverie", "👕", "#8b5cf6",
             "Machines disponibles 24h/24. Service blanchisserie (délai 24h). Casier personnel.",
             ["8 machines à laver", "4 sèche-linge", "Service pressing", "Lessive fournie"]),
            ("Sécurité", "🔒", "#64748b",
             "Badge obligatoire 24h/24. Rondes toutes les 2h. Caméras dans les espaces communs.",
             ["Contrôle accès 24h", "Caméras HD", "Équipe sécurité dédiée", "Coffre-fort réception"]),
            ("Lieu de culte", "🕌", "#ca8a04",
             "Espace de prière disponible bloc C. Moment de silence respecté par tous.",
             ["Accessible 24h", "Tapis fournis", "Orientation qibla", "Espace multi-confession"]),
            ("Transport", "🚌", "#0891b2",
             "Navettes camp ↔ mine matin et soir. Rotations Abidjan planifiées toutes les 2 semaines.",
             ["Navette mine 5h45 & 17h45", "Rotation Abidjan bi-mensuelle", "Réservation 72h à l'avance", "App mobile disponible"]),
        ]
        for ordre, (titre, emoji, couleur, desc, details) in enumerate(infras_data):
            InductionInfra.objects.create(
                titre=titre, emoji=emoji, couleur=couleur, description=desc,
                details=details, ordre=ordre, actif=True,
            )

        regles_data = [
            ("Tolérance Zéro Alcool", "🚫", "critique",
             "Toute consommation ou détention d'alcool est strictement interdite dans l'enceinte du camp. Violation = rapatriement immédiat sans préavis."),
            ("Couvre-feu Sonore 22h", "🔇", "important",
             "Silence obligatoire de 22h à 6h dans les résidences. Musique uniquement avec écouteurs. Respect du sommeil des collègues."),
            ("EPI Obligatoires", "👷", "critique",
             "Port du casque, gilet, lunettes et chaussures de sécurité obligatoire dans toutes les zones opérationnelles sans exception."),
            ("Contrôle des Accès", "🪪", "important",
             "Badge obligatoire 24h/24. Aucun visiteur sans autorisation écrite préalable de la direction. Tout accès non autorisé est signalé."),
            ("Économie d'Énergie", "⚡", "important",
             "Climatisation entre 20°C et 26°C uniquement. Lumières éteintes en quittant la chambre. Appareils énergivores (>100W) à déclarer."),
            ("Tri Sélectif Obligatoire", "♻️", "standard",
             "Bacs verts (organique), bleus (plastique/verre), noirs (ordures). Tri non respecté = pénalité sur bonus mensuel."),
            ("Usage Réseau", "📶", "standard",
             "Wi-Fi pour usage personnel raisonnable. Téléchargements massifs interdits. Streaming 4K limité aux heures creuses (22h–6h)."),
            ("Respect Mutuel", "🤝", "standard",
             "Langue inclusive, respect des différences culturelles et religieuses. Tout acte de harcèlement ou discrimination est un motif de licenciement."),
        ]
        for ordre, (titre, emoji, niveau, texte) in enumerate(regles_data):
            InductionRegle.objects.create(
                titre=titre, emoji=emoji, niveau=niveau, texte=texte,
                ordre=ordre, actif=True,
            )

        quiz_data = [
            ("À quelle heure commence le couvre-feu sonore dans les résidences ?",
             ["20h00", "21h00", "22h00", "00h00"], 2,
             "Le silence est obligatoire de 22h à 6h. Respecter le sommeil de vos collègues est essentiel."),
            ("Que devez-vous faire avec un appareil électrique de plus de 100W ?",
             ["Le garder discrètement", "Le déclarer à l'administration", "L'interdire totalement", "Ne rien faire"], 1,
             "Tout appareil de plus de 100W doit être déclaré pour la gestion énergétique du camp. La déclaration est gratuite."),
            ("En cas d'urgence médicale à 3h du matin, que faites-vous ?",
             ["Attendre le matin", "Appeler l'infirmerie (24h/24)", "Aller à la pharmacie", "Gérer seul"], 1,
             "L'infirmerie est ouverte 24h/24. N'hésitez jamais à appeler en cas d'urgence."),
            ("Peut-on inviter un ami à dormir au camp sans autorisation ?",
             ["Oui, entre amis", "Oui le week-end", "Absolument pas", "Oui si discret"], 2,
             "Aucun visiteur sans autorisation écrite de la direction. La sécurité du camp est l'affaire de tous."),
            ("Quelle température de climatisation est autorisée ?",
             ["Moins de 18°C", "Entre 20°C et 26°C", "N'importe quelle température", "Au-dessus de 28°C"], 1,
             "La plage 20-26°C est le compromis entre confort et économie d'énergie."),
        ]
        for ordre, (question, options, bonne_reponse, explication) in enumerate(quiz_data):
            InductionQuizQuestion.objects.create(
                question=question, options=options, bonne_reponse=bonne_reponse,
                explication=explication, ordre=ordre, actif=True,
            )

        return Response({
            "status": "ok",
            "infras_importees": len(infras_data),
            "regles_importees": len(regles_data),
            "quiz_importes": len(quiz_data),
        })

    @action(detail=False, methods=["get"])
    def actuelle(self, request):
        """Renvoie la configuration la plus récente, ou un objet vide si aucune n'existe.
        Auto-répare la table si elle n'existe pas encore (cas d'un déploiement où
        /api/setup-db/ n'a pas été appelé manuellement après l'ajout de ce modèle)."""
        default = {
            "id": None, "nom": "Camp Résidentiel", "site": "", "capacite": 0,
            "superficie": "", "altitude": "", "duree_parcours_min": 15,
        }
        try:
            cfg = InductionCampConfig.objects.order_by("-id").first()
        except Exception:
            _ensure_induction_tables()
            try:
                cfg = InductionCampConfig.objects.order_by("-id").first()
            except Exception:
                return Response(default)
        if not cfg:
            return Response(default)
        return Response(InductionCampConfigSerializer(cfg).data)


class InductionInfraViewSet(InductionAdminWriteMixin, viewsets.ModelViewSet):
    queryset = InductionInfra.objects.all()  # requis par DRF pour le basename du router
    serializer_class = InductionInfraSerializer

    def get_queryset(self):
        qs = InductionInfra.objects.all()
        if self.request.query_params.get("actives_only"):
            qs = qs.filter(actif=True)
        return qs.order_by("ordre", "id")


class InductionRegleViewSet(InductionAdminWriteMixin, viewsets.ModelViewSet):
    queryset = InductionRegle.objects.all()  # requis par DRF pour le basename du router
    serializer_class = InductionRegleSerializer

    def get_queryset(self):
        qs = InductionRegle.objects.all()
        if self.request.query_params.get("actives_only"):
            qs = qs.filter(actif=True)
        return qs.order_by("ordre", "id")


class InductionQuizQuestionViewSet(InductionAdminWriteMixin, viewsets.ModelViewSet):
    """Le serializer change selon le rôle : un admin voit la bonne réponse pour
    l'éditer, un agent qui passe le quiz ne la voit jamais dans la réponse réseau."""
    queryset = InductionQuizQuestion.objects.all().order_by("ordre", "id")

    def get_serializer_class(self):
        if _is_admin_user(self.request.user):
            return InductionQuizQuestionSerializer
        return InductionQuizQuestionPublicSerializer

    def get_queryset(self):
        qs = InductionQuizQuestion.objects.all()
        if self.request.query_params.get("actives_only"):
            qs = qs.filter(actif=True)
        return qs.order_by("ordre", "id")

    @action(detail=False, methods=["post"])
    def verifier(self, request):
        """Vérifie les réponses soumises sans jamais exposer les bonnes réponses
        d'avance. Payload attendu: {"reponses": {"<question_id>": <index_choisi>, ...}}"""
        reponses = request.data.get("reponses", {})
        questions = InductionQuizQuestion.objects.filter(actif=True)
        total = questions.count()
        correctes = 0
        detail = []
        for q in questions:
            choisi = reponses.get(str(q.id))
            ok = choisi is not None and int(choisi) == q.bonne_reponse
            if ok: correctes += 1
            detail.append({
                "id": q.id, "correct": ok,
                "bonne_reponse": q.bonne_reponse, "explication": q.explication,
            })
        score = round(correctes / total * 100) if total else 0
        return Response({"score": score, "correctes": correctes, "total": total, "detail": detail})
