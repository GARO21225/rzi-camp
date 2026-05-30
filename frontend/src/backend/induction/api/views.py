"""
API Views — Induction Module
ViewSets complets + endpoints métier
"""
from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from django.db.models import Q, Count
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend

from induction.models import (
    Site, Camp, Zone, Contractor,
    Employee, EmergencyContact,
    EmployeeDocument, DocumentType,
    Training, TrainingModule, EmployeeTraining,
    QuizQuestion, QuizChoice, QuizAttempt, QuizAnswer,
    MedicalCheck, AccessBadge, AccessLog,
    InductionWorkflow, WorkflowEvent,
)
from induction.serializers import (
    SiteSerializer, CampSerializer, ZoneSerializer, ContractorSerializer,
    EmployeeListSerializer, EmployeeDetailSerializer, EmployeeCreateSerializer,
    EmergencyContactSerializer,
    DocumentTypeSerializer, EmployeeDocumentSerializer, DocumentValidationSerializer,
    TrainingSerializer, TrainingModuleSerializer, EmployeeTrainingSerializer,
    QuizQuestionSerializer, QuizAttemptSerializer, QuizSubmitSerializer,
    MedicalCheckSerializer,
    AccessBadgeSerializer, AccessLogSerializer, QRScanSerializer,
    InductionWorkflowSerializer, DashboardSerializer,
)
from induction.permissions import (
    IsInductionAdmin, IsHRAgent, IsMedicalAgent, IsAccessAgent,
    IsSiteSupervisor, IsEmployeeOwner, IsAdminOrReadOnly,
)
from induction.services import (
    WorkflowService, BadgeService, AccessControlService, StatisticsService,
)


# ── Sites & Infrastructure ───────────────────────────

class SiteViewSet(viewsets.ModelViewSet):
    queryset            = Site.objects.filter(actif=True).prefetch_related('superviseurs')
    serializer_class    = SiteSerializer
    permission_classes  = [IsAuthenticated, IsAdminOrReadOnly]
    filter_backends     = [filters.SearchFilter, DjangoFilterBackend]
    search_fields       = ['nom', 'code', 'pays']
    filterset_fields    = ['actif', 'pays']

    @action(detail=True, methods=['get'])
    def dashboard(self, request, pk=None):
        """KPIs du site."""
        site = self.get_object()
        data = StatisticsService.dashboard(site_id=site.id)
        return Response(data)

    @action(detail=True, methods=['get'])
    def employes(self, request, pk=None):
        site = self.get_object()
        emp_qs = site.employees.filter(actif=True)
        page = self.paginate_queryset(emp_qs)
        ser  = EmployeeListSerializer(page or emp_qs, many=True)
        return self.get_paginated_response(ser.data) if page else Response(ser.data)


class CampViewSet(viewsets.ModelViewSet):
    queryset           = Camp.objects.filter(actif=True).select_related('site')
    serializer_class   = CampSerializer
    permission_classes = [IsAuthenticated, IsAdminOrReadOnly]
    filterset_fields   = ['site']


class ZoneViewSet(viewsets.ModelViewSet):
    queryset           = Zone.objects.filter(actif=True).select_related('site')
    serializer_class   = ZoneSerializer
    permission_classes = [IsAuthenticated, IsAdminOrReadOnly]
    filterset_fields   = ['site', 'type_zone', 'induction_requise']


class ContractorViewSet(viewsets.ModelViewSet):
    queryset           = Contractor.objects.filter(actif=True)
    serializer_class   = ContractorSerializer
    permission_classes = [IsAuthenticated, IsAdminOrReadOnly]
    search_fields      = ['nom', 'code']


# ── Employees ────────────────────────────────────────

class EmployeeViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    filter_backends    = [filters.SearchFilter, DjangoFilterBackend, filters.OrderingFilter]
    search_fields      = ['nom', 'prenom', 'email', 'matricule']
    filterset_fields   = ['site', 'type_employe', 'statut', 'contractor', 'actif']
    ordering_fields    = ['nom', 'prenom', 'created_at', 'statut']
    ordering           = ['nom']

    def get_queryset(self):
        return (Employee.objects
                .select_related('site', 'camp', 'contractor')
                .prefetch_related('contacts_urgence')
                .filter(actif=True))

    def get_serializer_class(self):
        if self.action == 'list':
            return EmployeeListSerializer
        if self.action in ('create', 'update', 'partial_update'):
            return EmployeeCreateSerializer
        return EmployeeDetailSerializer

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=['get'])
    def workflow(self, request, pk=None):
        employee = self.get_object()
        try:
            wf = employee.workflow
            return Response(InductionWorkflowSerializer(wf).data)
        except InductionWorkflow.DoesNotExist:
            return Response({'detail': 'Workflow non initialisé'}, status=404)

    @action(detail=True, methods=['post'])
    def demarrer_induction(self, request, pk=None):
        employee = self.get_object()
        site_id  = request.data.get('site_id')
        try:
            site = Site.objects.get(id=site_id) if site_id else employee.site
            if not site:
                return Response({'error': 'Site requis'}, status=400)
        except Site.DoesNotExist:
            return Response({'error': 'Site introuvable'}, status=404)
        wf = WorkflowService.initialiser(employee, site, request.user)
        return Response(InductionWorkflowSerializer(wf).data, status=201)

    @action(detail=True, methods=['get'])
    def mon_badge(self, request, pk=None):
        employee = self.get_object()
        badge = employee.badges.filter(statut='actif').order_by('-date_emission').first()
        if badge:
            return Response(AccessBadgeSerializer(badge).data)
        return Response({'detail': 'Aucun badge actif'}, status=404)

    @action(detail=True, methods=['get'])
    def historique_acces(self, request, pk=None):
        employee = self.get_object()
        logs = employee.historique_acces.select_related('site','zone').order_by('-timestamp')[:50]
        return Response(AccessLogSerializer(logs, many=True).data)


# ── Documents ────────────────────────────────────────

class DocumentTypeViewSet(viewsets.ModelViewSet):
    queryset           = DocumentType.objects.filter(actif=True)
    serializer_class   = DocumentTypeSerializer
    permission_classes = [IsAuthenticated, IsAdminOrReadOnly]


class EmployeeDocumentViewSet(viewsets.ModelViewSet):
    serializer_class   = EmployeeDocumentSerializer
    permission_classes = [IsAuthenticated]
    filter_backends    = [DjangoFilterBackend]
    filterset_fields   = ['employee', 'statut', 'type_doc']

    def get_queryset(self):
        return EmployeeDocument.objects.select_related(
            'employee', 'type_doc', 'valide_par'
        ).all()

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated, IsHRAgent])
    def valider(self, request, pk=None):
        doc = self.get_object()
        ser = DocumentValidationSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        action_val   = ser.validated_data['action']
        commentaire  = ser.validated_data.get('commentaire', '')
        if action_val == 'valider':
            doc.valider(request.user, commentaire)
            # Vérifier si tous les docs sont OK → avancer workflow
            try:
                ok, msg = WorkflowService.verifier_documents(doc.employee.workflow, request.user)
            except Exception:
                pass
            return Response({'status': 'validé', 'message': 'Document validé'})
        else:
            doc.refuser(request.user, commentaire)
            from induction.services import NotificationService
            NotificationService.document_refuse(doc.employee, doc, commentaire)
            return Response({'status': 'refusé', 'message': 'Document refusé'})

    @action(detail=False, methods=['get'])
    def en_attente(self, request):
        """Documents en attente de validation RH."""
        docs = EmployeeDocument.objects.filter(statut='soumis').select_related('employee','type_doc')
        return Response(EmployeeDocumentSerializer(docs, many=True).data)

    @action(detail=False, methods=['get'])
    def expirant_30j(self, request):
        import datetime
        limit = timezone.now().date() + datetime.timedelta(days=30)
        docs = (EmployeeDocument.objects
                .filter(statut='valide', date_expiration__lte=limit,
                        date_expiration__gte=timezone.now().date())
                .select_related('employee','type_doc'))
        return Response(EmployeeDocumentSerializer(docs, many=True).data)


# ── Formation ────────────────────────────────────────

class TrainingViewSet(viewsets.ModelViewSet):
    queryset           = Training.objects.filter(actif=True).prefetch_related('modules','sites')
    serializer_class   = TrainingSerializer
    permission_classes = [IsAuthenticated, IsAdminOrReadOnly]
    filter_backends    = [filters.SearchFilter, DjangoFilterBackend]
    search_fields      = ['titre', 'code']
    filterset_fields   = ['sites', 'obligatoire', 'type_formation']


class EmployeeTrainingViewSet(viewsets.ModelViewSet):
    serializer_class   = EmployeeTrainingSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields   = ['employee', 'formation', 'statut']

    def get_queryset(self):
        return EmployeeTraining.objects.select_related('employee','formation')

    @action(detail=True, methods=['post'])
    def completer_module(self, request, pk=None):
        emp_training = self.get_object()
        module_id = request.data.get('module_id')
        try:
            module = TrainingModule.objects.get(id=module_id, formation=emp_training.formation)
        except TrainingModule.DoesNotExist:
            return Response({'error': 'Module introuvable'}, status=404)
        emp_training.completer_module(module)
        if emp_training.statut == 'complete':
            try:
                ok, msg = WorkflowService.valider_formation(
                    emp_training.employee.workflow, emp_training.formation, request.user
                )
            except Exception:
                pass
        return Response(EmployeeTrainingSerializer(emp_training).data)


# ── Quiz ─────────────────────────────────────────────

class QuizQuestionViewSet(viewsets.ModelViewSet):
    queryset           = QuizQuestion.objects.filter(actif=True).prefetch_related('choix')
    serializer_class   = QuizQuestionSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields   = ['formation', 'niveau', 'sites']

    def get_queryset(self):
        import random
        qs = super().get_queryset()
        # Shuffle pour éviter l'ordre prévisible
        ids = list(qs.values_list('id', flat=True))
        random.shuffle(ids)
        preserved = {v:i for i,v in enumerate(ids)}
        return sorted(qs, key=lambda obj: preserved.get(obj.id, 0))


class QuizAttemptViewSet(viewsets.ModelViewSet):
    serializer_class   = QuizAttemptSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields   = ['employee', 'formation', 'statut', 'site']

    def get_queryset(self):
        return QuizAttempt.objects.select_related('employee','formation','site')

    @action(detail=False, methods=['post'])
    def demarrer(self, request):
        """Démarrer une nouvelle tentative de quiz."""
        employee_id  = request.data.get('employee_id')
        formation_id = request.data.get('formation_id')
        site_id      = request.data.get('site_id')
        try:
            employee  = Employee.objects.get(id=employee_id)
            formation = Training.objects.get(id=formation_id)
        except Exception:
            return Response({'error': 'Employé ou formation introuvable'}, status=404)
        # Compter les tentatives précédentes
        nb_tentatives = QuizAttempt.objects.filter(
            employee=employee, formation=formation
        ).count()
        tentative = QuizAttempt.objects.create(
            employee=employee, formation=formation,
            site_id=site_id,
            numero_tentative=nb_tentatives + 1,
            score_minimum=request.data.get('score_minimum', 80),
            supervise_par=request.user,
        )
        # Retourner les questions mélangées
        questions = QuizQuestion.objects.filter(
            formation=formation, actif=True
        ).prefetch_related('choix').order_by('?')[:20]  # Max 20 questions
        return Response({
            'tentative_id': str(tentative.id),
            'questions': QuizQuestionSerializer(questions, many=True).data,
            'score_minimum': tentative.score_minimum,
            'numero_tentative': tentative.numero_tentative,
        })

    @action(detail=True, methods=['post'])
    def soumettre(self, request, pk=None):
        """Soumettre les réponses et calculer le score."""
        tentative = self.get_object()
        if tentative.statut != 'en_cours':
            return Response({'error': 'Tentative déjà soumise'}, status=400)
        ser = QuizSubmitSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        # Enregistrer les réponses
        for rep in ser.validated_data['reponses']:
            try:
                question = QuizQuestion.objects.get(id=rep['question_id'])
                choix    = QuizChoice.objects.get(id=rep['choix_id'], question=question)
                QuizAnswer.objects.get_or_create(
                    tentative=tentative, question=question,
                    defaults={'choix_selectionne': choix}
                )
            except Exception:
                continue
        # Calculer score
        score = tentative.calculer_score()
        # Mettre à jour workflow
        try:
            wf = tentative.employee.workflow
            ok, msg = WorkflowService.valider_quiz(wf, tentative, request.user)
        except Exception:
            pass
        return Response({
            'score': float(score),
            'reussi': tentative.reussi,
            'score_minimum': tentative.score_minimum,
            'points': f"{tentative.points_obtenus}/{tentative.points_total}",
            'message': '🎉 Quiz réussi !' if tentative.reussi else '❌ Score insuffisant',
        })


# ── Médical ──────────────────────────────────────────

class MedicalCheckViewSet(viewsets.ModelViewSet):
    serializer_class   = MedicalCheckSerializer
    permission_classes = [IsAuthenticated, IsMedicalAgent]
    filter_backends    = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields   = ['employee', 'resultat', 'site']
    ordering           = ['-date_examen']

    def get_queryset(self):
        return MedicalCheck.objects.select_related('employee', 'site')

    def perform_create(self, serializer):
        visite = serializer.save(created_by=self.request.user)
        # Mettre à jour workflow
        try:
            wf = visite.employee.workflow
            WorkflowService.valider_medical(wf, visite, self.request.user)
            if wf.peut_valider:
                WorkflowService.valider_finale(wf, self.request.user)
        except Exception:
            pass

    @action(detail=False, methods=['get'])
    def expirant(self, request):
        import datetime
        limit = timezone.now().date() + datetime.timedelta(days=30)
        visites = (MedicalCheck.objects
                   .filter(resultat='FIT', date_expiration__lte=limit,
                           date_expiration__gte=timezone.now().date())
                   .select_related('employee'))
        return Response(MedicalCheckSerializer(visites, many=True).data)


# ── Contrôle d'accès ─────────────────────────────────

class AccessBadgeViewSet(viewsets.ModelViewSet):
    serializer_class   = AccessBadgeSerializer
    permission_classes = [IsAuthenticated]
    filter_backends    = [DjangoFilterBackend]
    filterset_fields   = ['employee', 'site', 'statut']

    def get_queryset(self):
        return AccessBadge.objects.select_related('employee','site').prefetch_related('zones_autorisees')

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated, IsInductionAdmin])
    def revoquer(self, request, pk=None):
        badge  = self.get_object()
        raison = request.data.get('raison', '')
        if not raison:
            return Response({'error': 'Raison de révocation requise'}, status=400)
        badge.revoquer(request.user, raison)
        # Mettre à jour statut employee
        badge.employee.statut = 'suspendu'
        badge.employee.save(update_fields=['statut'])
        return Response({'status': 'révoqué', 'message': f'Badge {badge.qr_code_string} révoqué'})

    @action(detail=True, methods=['get'])
    def telecharger_pdf(self, request, pk=None):
        """Télécharger le badge PDF."""
        from django.http import HttpResponse
        badge = self.get_object()
        if not badge.badge_base64:
            badge.badge_base64 = BadgeService.generer_badge_base64(badge)
            badge.save(update_fields=['badge_base64'])
        import base64
        pdf_data = base64.b64decode(badge.badge_base64)
        response = HttpResponse(pdf_data, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="badge_{badge.qr_code_string}.pdf"'
        return response


class AccessLogViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class   = AccessLogSerializer
    permission_classes = [IsAuthenticated, IsAccessAgent]
    filter_backends    = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields   = ['employee', 'site', 'zone', 'resultat']
    ordering           = ['-timestamp']

    def get_queryset(self):
        return (AccessLog.objects
                .select_related('employee','site','zone','badge','agent_scan')
                .all())


class QRScanView(APIView):
    """Endpoint principal de scan QR — contrôle d'accès."""
    permission_classes = [IsAuthenticated, IsAccessAgent]

    def post(self, request, site_id):
        ser = QRScanSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        result = AccessControlService.verifier_qr(
            qr_string=ser.validated_data['qr_string'],
            site_id=site_id,
            zone_id=ser.validated_data.get('zone_id'),
            agent=request.user,
        )
        http_status = status.HTTP_200_OK if result['autorise'] else status.HTTP_403_FORBIDDEN
        return Response(result, status=http_status)


# ── Workflow ──────────────────────────────────────────

class InductionWorkflowViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class   = InductionWorkflowSerializer
    permission_classes = [IsAuthenticated]
    filter_backends    = [DjangoFilterBackend]
    filterset_fields   = ['employee', 'site', 'statut']

    def get_queryset(self):
        return (InductionWorkflow.objects
                .select_related('employee','site','badge')
                .prefetch_related('events')
                .all())

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated, IsInductionAdmin])
    def valider(self, request, pk=None):
        wf = self.get_object()
        if not wf.peut_valider:
            return Response({'error': 'Toutes les étapes ne sont pas complètes',
                             'progression': wf.progression_pct}, status=400)
        ok, msg = WorkflowService.valider_finale(wf, request.user)
        if ok:
            return Response({'status': 'validé', 'message': msg,
                             'badge': wf.badge.qr_code_string if wf.badge else None})
        return Response({'error': msg}, status=400)

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated, IsInductionAdmin])
    def refuser(self, request, pk=None):
        wf     = self.get_object()
        raison = request.data.get('raison', '')
        wf.statut = 'refuse'
        wf.raison_refus = raison
        wf.save(update_fields=['statut','raison_refus'])
        wf.employee.statut = 'refuse'
        wf.employee.save(update_fields=['statut'])
        WorkflowEvent.objects.create(
            workflow=wf, action='workflow_reset',
            description=f'Refusé: {raison}', effectue_par=request.user
        )
        return Response({'status': 'refusé', 'raison': raison})


# ── Dashboard Global ──────────────────────────────────

class DashboardView(APIView):
    permission_classes = [IsAuthenticated, IsHRAgent]

    def get(self, request):
        site_id = request.query_params.get('site_id')
        data    = StatisticsService.dashboard(site_id=site_id)
        return Response(data)


class ExportView(APIView):
    """Export Excel des données d'induction."""
    permission_classes = [IsAuthenticated, IsInductionAdmin]

    def get(self, request, export_type):
        from django.http import HttpResponse
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return Response({'error': 'openpyxl non installé'}, status=500)

        wb = openpyxl.Workbook()
        ws = wb.active
        header_style = Font(bold=True, color='FFFFFF')
        header_fill  = PatternFill(fgColor='1E3A8A', fill_type='solid')

        if export_type == 'employes':
            ws.title = 'Employés'
            headers = ['Nom','Prénom','Email','Type','Statut','Site','Société','Date arrivée']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font, cell.fill = header_style, header_fill
            for row, emp in enumerate(Employee.objects.select_related('site','contractor').filter(actif=True), 2):
                ws.append([emp.nom, emp.prenom, emp.email,
                           emp.get_type_employe_display(),
                           emp.get_statut_display(),
                           emp.site.nom if emp.site else '',
                           emp.contractor.nom if emp.contractor else '',
                           str(emp.date_arrivee or '')])

        elif export_type == 'badges':
            ws.title = 'Badges actifs'
            headers = ['Employé','Site','Badge N°','Date émission','Date expiration','Statut']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font, cell.fill = header_style, header_fill
            for row, badge in enumerate(AccessBadge.objects.select_related('employee','site').filter(statut='actif'), 2):
                ws.append([badge.employee.nom_complet, badge.site.nom, badge.qr_code_string,
                           str(badge.date_emission), str(badge.date_expiration), badge.statut])

        elif export_type == 'conformite':
            ws.title = 'Conformité'
            headers = ['Employé','Email','Type','Statut','Documents OK','Quiz OK','Médical OK','Progression']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font, cell.fill = header_style, header_fill
            for row, emp in enumerate(Employee.objects.filter(actif=True).select_related('site'), 2):
                try:
                    wf = emp.workflow
                    ws.append([emp.nom_complet, emp.email, emp.get_type_employe_display(),
                               emp.get_statut_display(),
                               '✅' if wf.etape_documents else '❌',
                               '✅' if wf.etape_quiz else '❌',
                               '✅' if wf.etape_medical else '❌',
                               f"{wf.progression_pct}%"])
                except Exception:
                    ws.append([emp.nom_complet, emp.email, '', emp.get_statut_display(),
                               '❌','❌','❌','0%'])

        buf = __import__('io').BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="induction_{export_type}.xlsx"'
        return response
